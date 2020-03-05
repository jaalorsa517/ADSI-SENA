# -*- coding:utf-8 -*-
'''
Este script va a recopilar todos los clientes de todos los inventarios por ciudad
'''

import openpyxl
import sqlite3
from openpyxl import load_workbook


class getClientes:
    def __init__(self, id='', nom='', neg='', tel='', email='', dir=''):
        self.id = id
        self.nom = nom
        self.neg = neg
        self.tel = tel
        self.email = email
        self.dir = dir


def load():

    print('INICIO PROCESO CLIENTES')

    _PATH_DB = '/home/jaalorsa/Proyectos/flutter/ventas/data/pedidoDB.db'
    _PATH_XLXS = '/home/jaalorsa/Documentos/Drive/Gralac/'
    _PATHS_XLSX = ('Andes/AndesInventario.xlsx',
                   'Betania/BetaniaInventario.xlsx',
                   'Hispania/HispaniaInventario.xlsx',
                   'Jardin/JardinInventario.xlsx',
                   'Santa_Ines/SantaInesInventario.xlsx',
                   'Santa_Rita/SantaRitaInventario.xlsx',
                   'Taparto/TapartoInventario.xlsx')

    try:
        con = sqlite3.connect(_PATH_DB)
        cursor = con.cursor()

        cursor.execute("DELETE FROM cliente")
        cursor.execute("DELETE FROM 'ciudad-cliente'")
        con.commit()

    except sqlite3.Error as e:
        print(type(e).__name__)

    finally:
        con.close()

    for xlsx in _PATHS_XLSX:
        _workbook = load_workbook(_PATH_XLXS + xlsx)
        dic = {}

        for sheet in _workbook.sheetnames:
            dic.clear()
            a1 = _workbook[sheet]['A1']
            _c = getClientes()

            if (a1.value is not None):

                if (a1.comment is not None and len(a1.comment.text) != 0):
                    dic = datosInComment(a1.comment.text)

                    if ('nom' in dic):

                        if (not dic['nom'] or dic.get('nom').isspace()):
                            dic['nom'] = a1.value
                    else:
                        dic['nom'] = a1.value

                    dic['neg'] = a1.value

                else:
                    dic['nom'] = a1.value
                    dic['neg'] = a1.value

                if ('nom' in dic):
                    _c.nom = dic['nom']
                if ('id' in dic):
                    _c.nom = dic['id']
                if ('neg' in dic):
                    _c.neg = dic['neg']
                if ('tel' in dic):
                    _c.tel = dic['tel']
                if ('email' in dic):
                    _c.email = dic['email']

                try:
                    con = sqlite3.connect(_PATH_DB)
                    cursor = con.cursor()

                    cursor.execute(
                        "INSERT INTO cliente (id,nombre,representante,telefono,email,direccion) VALUES('{}','{}','{}','{}','{}','{}')"
                        .format(_c.id, _c.neg, _c.nom, _c.tel, _c.email,
                                _c.dir))

                    cursor.execute(
                        "SELECT id1 FROM cliente WHERE nombre='{}'".format(
                            _c.neg))
                    var = cursor.fetchone()

                    cursor.execute(
                        "INSERT INTO 'ciudad-cliente' VALUES ({},{});".format(
                            var[0], (_PATHS_XLSX.index(xlsx) + 1)))

                    con.commit()

                except sqlite3.Error as e:
                    print(type(e).__name__)

                finally:
                    con.close()
        print('ARCHIVO TERMINADO')
    print('FIN PROCESO CLIENTES')

def datosInComment(coment):
    '''
    Funcion que separa un texto con el patron'\\n' y a
    su vez se vuleve a separar con el patron ':'
    _datos=coment.split('\\n')
    @param coment: String con el patron 
                    id:123\\nnom:nom\\ntel:123\\email:email
    @return dic{cc:,nom:,tel:,email}
    '''
    _datos = coment.split('\n')
    datos_dic = {}
    if (len(_datos) != 0):
        for d in _datos:
            i = d.split(':')
            if (len(i) == 2):
                datos_dic[i[0]] = i[1]
    return datos_dic